from openpyxl import load_workbook
from .models import Course


def get_xlsx():
    workbook = load_workbook(filename='路径/文件名.xlsx')
    sheet = workbook['工作表名']
    # 读取指定行和列

    sheet = workbook['工作表名']

    # 遍历表格的所有行
    for row in sheet.iter_rows(values_only=True):
        # 遍历行中的每个单元格
        count = 0
        for cell_value in row:
            count += 1
            if count == 1:
                name = cell_value
            elif count == 4:
                department = cell_value
            elif count == 6:
                teacher_name = cell_value
            elif count == 9:
                id = cell_value
        Course.objects.create(id=id, name=name, department=department, teacher_name=teacher_name)

# A1 名称  B2 本科生/研究生 C3 校区 D4 开课学院 E5 课程类别  F6老师姓名 G7课程性质（必修等）H8节次
# I9 课程代码 J10 选课人数 K11 老师职称 L12 上课时间 M13 上课地点
#
