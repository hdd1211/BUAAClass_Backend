# Create your views here.

import random
import random
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from wordcloud import WordCloud
from matplotlib import pyplot as plt
from .models import Course
from django.http import FileResponse
from openpyxl import load_workbook

from django.http import HttpResponse, JsonResponse
# from .utils.response import wrap_response_data
from .models import User, Evaluation, Course, Report, Admin
from .serilization import LoginSerializer, RegisterSerializer, InteractionsSerializer, CourselistSerializer, \
    AnnouncementSerilizer, EvaluationSerilizer, AdminLoginSerializer, CourseSerializer, DelcourseSerializer
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from functools import wraps
from django.http import JsonResponse
from rest_framework.decorators import api_view


# from .read import get_xlsx


def get_xlsx():
    print(1)
    workbook = load_workbook(filename='F:\Project\Python\djangoProject1\class.xlsx')
    print(333)
    sheet = workbook['工作表1']
    print(2)
    # 遍历表格的所有行
    for row in sheet.iter_rows(values_only=True):
        # 遍历行中的每个单元格
        print('1')
        count = 0
        for cell_value in row:
            count += 1
            if count == 1:
                name = cell_value
            elif count == 4:
                department = cell_value
            elif count == 6:
                teacher_name = cell_value
            elif count == 9:
                id = cell_value

        # 检查课程是否已经存在
        course, created = Course.objects.get_or_create(
            id=id,
            defaults={'name': name, 'department': department, 'teacher_name': teacher_name}
        )
        if created:
            print(f'Created new course: {course.id}')


# from .admin import app_admin
def wordcloud(request):
    # 首先，我们从数据库中随机选择十门课程
    courses = list(Course.objects.all())
    selected_courses = random.sample(courses, min(10, len(courses)))

    # 然后，我们创建一个包含这些课程名的字符串
    course_names = " ".join(course.name for course in selected_courses)

    # 接下来，我们使用这个字符串创建一个词云
    wordcloud = WordCloud(width=800, height=400).generate(course_names)

    # 将词云保存为PNG图片
    img_temp = NamedTemporaryFile(delete=True)
    img_temp.write(wordcloud.to_image().tobytes())
    img_temp.flush()

    return FileResponse(open(img_temp.name, 'rb'), as_attachment=True, filename='course_wordcloud.png')


def index(request):
    return HttpResponse("Hello, world. You're at the eva index.")


def custom_login_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'status': 1, 'message': 'Invalid login'}, status=401)
        return view_func(request, *args, **kwargs)

    return wrapper


def admin_login_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.admin.is_authenticated:
            return JsonResponse({'status': 1, 'message': 'Invalid login'}, status=401)
        return view_func(request, *args, **kwargs)

    return wrapper


@api_view()
def login(request):
    if request.method == 'POST':
        '''111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111'''
        user_name = request.POST.get('user_name')
        password = request.POST.get('password')
        user = User.objects.get(user_name=user_name)
        if User.objects.filter(user_name=user_name):
            if User.objects.filter(password=password):
                response_data = LoginSerializer.get_login(0, user)
            else:
                response_data = LoginSerializer.get_login(1, user)
        else:
            response_data = LoginSerializer.get_login(1, user)

    response_data['agree_reviews'] = None
    response_data['disagree_reviews'] = None
    response_data['reviews'] = None
    return JsonResponse(response_data)


@api_view()
def register(request):
    if request.method == 'POST':
        user_name = request.POST.get('user_name')
        password = request.POST.get('password')
        if User.objects.filter(user_name=user_name).exists():
            reponse_data = RegisterSerializer.get_register(1)
        elif len(password) < 6 or len(password) > 16 or len(user_name) > 20:
            response_data = RegisterSerializer.get_register(2)
        else:
            user = User.objects.create(user_name=user_name, password=password)
            reponse_data = RegisterSerializer.get_register(0)
            # 如果有name和password不符合规定的情况怎么处理？#######################################################
            # 已解决
    return JsonResponse(reponse_data)


@api_view()
def review(request):
    if request.method == 'POST':
        id = request.POST.get(course_id)


@custom_login_required
@api_view()
def release(request):
    if request.method == 'POST':
        evaluation = Evaluation.objects.create(user_id=request.POST.get('user_id'),
                                               smester=request.POST.get('smester'),
                                               rating_total=request.POST.get('smester'),
                                               rating_quality=request.POST.get('rating_quality'),
                                               rating_workload=request.POST.get('rating_workload'),
                                               rating_assesment=request.POST.get('rating_assesment'),
                                               title=request.POST.get('title'),
                                               text=request.POST.get('text')
                                               )
        return 0


@custom_login_required
@api_view()
def interaction(request):
    if request.method == 'POST':
        review_id = request.POST.get('review_id')
        op = request.POST.get('op')

        if op == 0:
            evaluation = Evaluation.objects.filter(id(review_id)).update(agree_cnt=agree_cnt + 1)
        elif op == 1:
            evaluation = Evaluation.objects.filter(id(review_id)).update(disagree_cnt=disagree_cnt + 1)

        response_data = InteractionsSerializer.get_interaction(0, evaluation.get_agree_cnt(evaluation))
    return JsonResponse(response_data)


@api_view(['GET'])
def course_list(request):
    if request.method == 'GET':
        get_xlsx()
        course_data = get_courselist()  # 获取课程列表数据
        serializer = CourselistSerializer(course_data, many=True)  # 创建序列化器实例
        return Response(serializer.data)  # 返回序列化的数据
    return JsonResponse("response_data")


@custom_login_required
@api_view(['POST'])
def report(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        report_text = request.POST.get('report')

        report = Report.objects.create(user_id=user_id, content=report_text)
        return 0


@custom_login_required
@api_view()
def concerning(request):
    if request.method == 'POST':
        op = request.POST.get('op')
        if op == 1:
            return 0000000


@custom_login_required
@api_view()
def recommend(request):
    if request.method == 'POST':
        evaluations = Evaluation.objects.all().ordered_by('-agree_cnt', '-disagree_cnt')
        size = Evaluation.objects.count() / 10
        evaluation = Evaluation.objects.get(random.randint(0, size))
        response_data = EvaluationSerilizer.get_evaluation(evaluation.id)
        return JsonResponse(response_data)


@custom_login_required
@api_view()
def announcement(request):
    if request.method == 'POST':
        response_data = AnnouncementSerilizer.get_announcement()
        return JsonResponse(response_data)


@custom_login_required
@api_view()
def draft(request):
    1


@api_view()
def admin_login(request):
    if request.method == 'POST':
        '''111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111111'''
        user_name = request.POST.get('admin_name')
        password = request.POST.get('password')
        admin = Admin.objects.get(admin_name=admin_name)
        if Admin.objects.filter(admin_name=admin_name):
            if Admin.objects.filter(password=password):
                response_data = AdminLoginSerializer.get_login(0, admin)
            else:
                response_data = AdminLoginSerializer.get_login(1, admin)
        else:
            response_data = AdminLoginSerializer.get_login(1, admin)
    return JsonResponse(response_data)


@admin_login_required
@api_view()
def admin_catalog_all(request):
    if request.method == 'POST':
        courselist(request)


@admin_login_required
@api_view()
def admin_catalog_get_by_name(request):
    if request.method == 'POST':
        name = request.POST.get(name)
        courses = Course.objects.filter(name=name)
        response_data = CourselistSerializer(courses)
        return response_data


@admin_login_required
@api_view()
def admin_catalog_get_by_id(request):
    if request.method == 'POST':
        id = request.POST.get(id)
        try:
            course = Course.objects.get(id=id)
            response_data = CourseSerializer(course)
            return response_data
        except Course.DoesNotExist:
            model_data = model_to_dict(Course)
            response_data = dict.fromkeys(model_data.keys(), None)
            return response_data


@admin_login_required
@api_view()
def admin_catalog_get_by_type(request):
    if request.method == 'POST':
        type = request.POST.get(type)
        try:
            courses = Course.objects.filter(type=type)
            response_data = CourselistSerializer(courses)
            return response_data
        except Course.DoesNotExist:
            model_data = model_to_dict(Course)
            response_data = dict.fromkeys(model_data.keys(), None)
            return response_data


@admin_login_required
@api_view()
def admin_catalog_del_course(request):
    if request.method == 'POST':
        id = request.POST.get(id)
        try:
            course = Course.objects.get(id=id)
            course.delete()
            response_data = DelcourseSerializer(0)
            return response_data
        except Course.DoesNotExist:
            response_data = DelcourseSerializer(1)
            return response_data


@admin_login_required
@api_view()
def admin_catalog_del_batch(request):
    if request.method == 'POST':
        ids = request.POST.getlist(id_batch)
        i = 0
        for id in ids:
            admin_catalog_del_course(request.POST.get[i])
            i += 1


@admin_login_required
@api_view()
def admin_catalog_add_batch(request):
    1


@api_view(['GET'])
def get_all_reviews(request):
    # 从数据库中获取所有评论
    reviews = Review.objects.all()
    # 使用序列化器将评论对象转换为JSON格式
    serializer = ReviewSerializer(reviews, many=True)
    # 将结果返回给客户端
    return Response({"review_list": serializer.data})


@api_view(['GET'])
def get_reviews_by_status(request, status):
    # 从数据库中筛选出特定状态的评论
    reviews = Evaluation.objects.filter(status=status)
    # 使用序列化器将评论对象转换为JSON格式
    serializer = ReviewSerializer(reviews, many=True)
    # 将结果返回给客户端
    return Response({"review_list": serializer.data})


@api_view(['GET'])
def get_reviews_by_user(request, user_id):
    # 从数据库中筛选出由特定用户发表的评论
    reviews = Review.objects.filter(user_id=user_id)
    # 使用序列化器将评论对象转换为JSON格式
    serializer = ReviewSerializer(reviews, many=True)
    # 将结果返回给客户端
    return Response({"review_list": serializer.data})


@api_view(['GET'])
def get_reviews_by_course(request, course_id):
    # 从数据库中筛选出针对特定课程的评论
    reviews = Review.objects.filter(course_id=course_id)
    # 使用序列化器将评论对象转换为JSON格式
    serializer = ReviewSerializer(reviews, many=True)
    # 将结果返回给客户端
    return Response({"review_list": serializer.data})


@api_view(['GET'])
def get_reviews_by_word(request, word):
    # 使用Q对象进行复杂查询，筛选出评论标题或正文包含特定关键字的评论
    reviews = Review.objects.filter(Q(title__contains=word) | Q(content__contains=word))
    # 使用序列化器将评论对象转换为JSON格式
    serializer = ReviewSerializer(reviews, many=True)
    # 将结果返回给客户端
    return Response({"review_list": serializer.data})


@api_view(['DELETE'])
def del_review(request, id):
    try:
        # 获取指定id的评论
        review = Review.objects.get(id=id)
        # 将评论的状态改为已删除
        review.status = 2  # 假设已删除的状态码为2
        review.save()
        # 返回成功的结果
        return Response({"res": 0})
    except Review.DoesNotExist:
        # 如果指定id的评论不存在，返回失败的结果
        return Response({"res": 1})


@api_view(['DELETE'])
def del_batch(request):
    try:
        # 从请求体中获取要删除的评论id列表
        id_batch = request.data.get('id_batch', [])
        # 找到所有指定id的评论并修改它们的状态
        Review.objects.filter(id__in=id_batch).update(status=2)  # 假设已删除的状态码为2
        # 返回成功的结果
        return Response({"res": 0})
    except Exception as e:
        # 如果出错，返回失败的结果
        return Response({"res": 1})


@api_view(['GET'])
def get_review_by_id(request):
    # 获取URL参数中的id
    review_id = request.GET.get('id', None)

    if review_id is None:
        # 如果没有提供id，返回错误结果
        return Response({"error": "Missing 'id' parameter."}, status=400)

    try:
        # 根据id获取评论
        review = Review.objects.get(id=review_id)
        # 序列化评论对象
        serializer = ReviewSerializer(review)
        # 返回评论信息
        return Response(serializer.data)
    except Review.DoesNotExist:
        # 如果找不到评论，返回错误结果
        return Response({"error": "Review not found."}, status=404)
# @api_view()
# def add_course(request):
#     if request.method=='POST':
#         course =
# def delete_user(request):
#     user_name = request.POST.get('user_name')
#     if (User.objects.filter(user_name=user_name).exists()):
#         user = User.objects.get(user_name=user_name)
#         user.delete()
#         response_data = {'success': 0, 'message': '成功注销'}
#         return JsonResponse(response_data)
#     else:
#         response_data = {'success': 1, 'message': '账号不存在'}
#         return JsonResponse(response_data)
#
#
# # def get_user_needmodify(request):
# #     user_name = request.POST.get('user_name')
# #     if (User.objects.filter(user_name=user_name).exists()):
# #         user = User.objects.get(user_name=user_name)
# #
# #         return user
# #     else:
# #         return 1
#
#
# def modify_password(request):
#     user_name = request.POST.get('user_name')
#     new_password = request.POST.get('new_password')
#     if (User.objects.filter(user_name=user_name).exists()):
#         user = User.objects.get(user_name=user_name)
#         user.password = new_password
#         response_data = {'success': 0, 'message': '成功修改密码'}
#         return JsonResponse(response_data)
#     else:
#         response_data = {'success': 1, 'message': '未找到用户'}
#         response_data = wrap_response_data(0)
#         return JsonResponse(response_data)
#
